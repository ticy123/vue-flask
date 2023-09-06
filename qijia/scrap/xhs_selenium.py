import csv
import datetime
import threading
from time import sleep

import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
import os



os.system("taskkill /f /im chromedriver.exe")
chrome_options = Options()

chrome_options.add_argument(
    'user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.182 Safari/537.36",Connection="close"')
chrome_options.add_argument('blink-settings=imagesEnabled=false')  # 禁用图片
#加速
# chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_argument('--disable-extensions')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--disable-infobars')
chrome_options.add_argument('--disable-popup-blocking')
chrome_options.add_argument('--hide-scrollbars')
chrome_options.add_argument('--ignore-certificate-errors')
chrome_options.add_argument('--ignore-ssl-errors')
chrome_options.add_argument('--lang=zh-CN')
# chrome_options.add_argument('--log-level=3')
# chrome_options.add_argument('--no-proxy-server')
chrome_options.add_argument('--no-sandbox')
# chrome_options.add_argument('--start-maximized')
chrome_options.add_argument('--use-gl=swiftshader')
chrome_options.add_argument('--disable-software-rasterizer')
chrome_options.page_load_strategy = 'normal'


class Brower:

    def __init__(self, file_name, start_date='2023-05-01', end_date='2023-05-31'):
        self._download_path='C:\\Users\\Administrator\\Downloads'
        self._start_date = start_date
        self._end_date = end_date
        self._driver = webdriver.Chrome(options=chrome_options, executable_path='C:\chromedriver_win32\chromedriver')
        self._driver.set_page_load_timeout(10)
        self._driver.set_script_timeout(10)
        self._wait = WebDriverWait(self._driver, 10,2,ignored_exceptions=None)
        self._file_name = file_name
        self._items = []
        self._exist_account = ['-'] + self.read_excel()
        self._pages = 0

    def set_pages(self, value):
        self._pages = value

    def set_exist_account(self, name):
        self._exist_account.append(name)

    def set_items(self, account_name, total_account, return_account):
        self._items.append(
            {"account_name": account_name, "total_account": total_account, "return_account": return_account})

    def is_exist_account(self, account_name):
        return account_name in self._exist_account


    @property
    def download_path(self):
        return self._download_path

    @property
    def exist_account(self):
        return self._exist_account

    @property
    def start_date(self):
        return self._start_date

    @property
    def end_date(self):
        return self._end_date

    @property
    def pages(self):
        return self._pages

    @property
    def limit(self):
        return self._limit

    @property
    def driver(self):
        return self._driver

    @property
    def wait(self):
        return self._wait

    @property
    def file_name(self):
        return self._file_name

    @property
    def items(self):
        return self._items


    def login(self):
        count =0
        while True:
            try:
                print("开始登录")
                self.driver.maximize_window()
                self.driver.get("https://ad.xiaohongshu.com/aurora/home")
                actions = webdriver.ActionChains(self.driver)
                el = self.wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "css-1r2f04i")))
                actions.move_to_element(el[1]).click().perform()
                # print("点击登录按钮")
                self.wait.until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="邮箱"]'))).send_keys(
                    "qitiantian@qeeka.com")
                self.wait.until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="密码"]'))).send_keys(
                    "Qiyixhs@2108")
                self.wait.until(
                    EC.presence_of_element_located((By.XPATH, '//button[@class="css-r7neow css-wp7z9d dyn"]'))).click()
                print("登录完成")
                sleep(5)
                return True
            except Exception as e:
                count += 1
                if count == 3:
                    print("登录失败")
                    raise e
                print(f"login - {count}:{e}")
                self.driver.refresh()
                sleep(5)
                continue

    def get_all_accounts(self):
        try:
            ("点击工具")
            self.driver.find_element(By.XPATH,'/ html / body / div[1] / div / div[1] / div / div[1] / div[2] / div / div[4] / div / span[1]').click()
            sleep(2)
            print("点击子账户管理")
            self.driver.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/div[1]/div[1]/div[3]/div[2]/div[1]/span[2]').click()
            sleep(5)
            self.driver.execute_script("window.scrollTo(0 ,document.body.scrollHeight)")
            sleep(2)
            # 获取总页数
            total_page = self.wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'd-pagination-page-content')))[-1].text
            print(f"总页数为:{total_page}")
            self.set_pages(int(total_page))
            sleep(5)
            # 获取每一页的子账户名
        except Exception as e:
            print('get_all_accounts',e)
            self.driver.refresh()
            self.get_all_accounts()
            sleep(10)
        self.enter_child_account()

    def enter_child_account(self):
        print("获取子账户数据")
        count = len(self.exist_account)
        while True:
            el, account_name = self.get_current_account_el()
            if el =="error":
                continue
            if el is None:
                return
            print(f"\n进入账户 {count}:{account_name}")
            el.click()
            sleep(5)
            # 获取子账户金额信息,并写入excel
            total_account, return_account = self.get_child_account_info()
            sleep(2)
            # 下载各地消费信息
            status = self.download_city_consume_info()[0]
            total_count = [account_name, total_account, return_account]
            city_list = [] if status == 1001 else self.read_city_consume_info(account_name)
            self.write_to_excel(total_count, city_list)
            self.set_exist_account(account_name)
            #self.set_items(account_name, total_account, return_account)
            self.return_main_account()
            count = count + 1
            sleep(5)

    def get_current_account_el(self):
        count = 0
        while True:
            try:
                res = None
                account_name = None
                for page in range(1, int(self.pages) + 1):
                    els = self.wait.until(EC.presence_of_all_elements_located((By.LINK_TEXT, '进入账号')))
                    print(f"page:{page},数量{len(els)}")
                    # 进入每一个子账户,获取信息
                    for el in els:
                        parent_el = el.find_element(By.XPATH, '../..')
                        prev_el = parent_el.find_element(By.XPATH, 'preceding-sibling::div[position()=3]')
                        # 获取账户名
                        account_name = prev_el.text
                        if not self.is_exist_account(account_name):
                            res = el
                            break
                    if res is not None:
                        break
                    if page < self.pages:
                        # next_page = self.wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[2]/div[3]/div[2]/div[2]/div[3]/div[2]/div[1]/div[5]')))
                        next_page = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//span[@class="d-icon --color-current --color-static --size-icon-default"]')))[-1]
                        print(f"第{page}页解析结束,进入下一页")
                        next_page.click()
                        sleep(5)
                        #将滚轮移到操作那一列
                        el_scroll = self.wait.until(EC.presence_of_element_located((By.XPATH, '//span[text()="操作"]')))
                        self.driver.execute_script("arguments[0].scrollIntoView();", el_scroll)
                        sleep(1)
                return res, account_name
            except Exception as e:
                count += 1
                if count == 10:
                    print("获取当前账户失败")
                    return "error", "error"
                print(f"get_current_account_el : {e}")
                self.driver.refresh()
                sleep(2)
                continue


    def return_main_account(self):
        count = 0
        while True:
            try:
                actions = webdriver.ActionChains(self.driver)
                user_el = self.driver.find_elements(By.CLASS_NAME, "userName")[0]
                actions.move_to_element(user_el).click().perform()
                self.wait.until(
                    EC.presence_of_element_located((By.XPATH, '//span[text()=" 返回主账户 "]'))).click()
                sleep(3)
                return True
            except Exception as e:
                count = count + 1
                print(f"return_main_account - {count} : {e}")
                if count == 5:
                    print("返回主账户失败")
                    raise e
                self.driver.refresh()
                sleep(3)
                continue

    def get_child_account_info(self):
        #失败重试10次
        count = 0
        while True:
            try:
                self.wait.until(EC.presence_of_element_located((By.XPATH, '//span[text()="财务"]'))).click()
                sleep(2)
                # 将self.start_date 和 self.end_date加一天
                start_date = (datetime.datetime.strptime(self.start_date, "%Y-%m-%d") + datetime.timedelta(
                    days=1)).strftime("%Y-%m-%d")
                end_date = (datetime.datetime.strptime(self.end_date, "%Y-%m-%d") + datetime.timedelta(
                    days=1)).strftime("%Y-%m-%d")
                start_date_input = self.wait.until(EC.presence_of_all_elements_located((By.XPATH,"//input[@class='d-text']")))[0]
                start_date_input.send_keys(Keys.CONTROL, 'a')
                start_date_input.send_keys(start_date)
                sleep(2)
                end_date_input = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, "//input[@class='d-text']")))[1]
                end_date_input.send_keys(Keys.CONTROL, 'a')
                end_date_input.send_keys(end_date)
                sleep(3)
                #查找元素，其中style里面包含‘2/2/3/3’的元素
                # 获取现金金额
                total_account = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='d-grid-item d-td d-table-cell-border-bottom' and contains(@style,'2 / 3 / 3 / 4')]"))).text
                # 获取返回金额
                return_account = self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='d-grid-item d-td d-table-cell-border-bottom' and contains(@style,'2 / 4 / 3 / 5')]"))).text
                total_account = total_account if total_account != '-' else 0
                return_account = return_account if return_account != '-' else 0
                print(f"总金额:{total_account},返现金额:{return_account}")
                return total_account, return_account
            except Exception as e:
                count = count + 1
                print(f"get_child_account_info - {count} : {e}")
                if count ==5 :
                    return "error","error"
                self.driver.refresh()
                sleep(5)
                continue


    # 开启一个新线程,处理随机弹窗
    def handle_alert(self):
        t = threading.Thread(target=self.alert)
        t.setDaemon(True)
        t.start()

    def alert(self):
        # 当随机弹窗存在时，将它关闭
        while True:
            try:
                self.driver.find_element(By.XPATH,"//span[text()='已读']").click()
                self.driver.switch_to.alert.accept()
                sleep(1)
            except:
                sleep(1)
                continue

    # @retry(stop_max_attempt_number=2, wait_fixed=3000)
    def start(self):
        self.login()
        self.get_all_accounts()

    def quit(self):
        self.driver.close()  # 关闭当前窗口
        self.driver.quit()  # 关闭所有窗口

    def read_excel(self):
        file_name = self.file_name
        data_list = []
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            rows = ws.rows
            data_list = []
            for row in rows:
                row_list = []
                for cell in row:
                    row_list.append(cell.value)
                data_list.append(row_list)
            data_list = [i[0] for i in data_list if i[0] != 'name']
            print(f'Excel已读取数据，为{data_list}')
        else:
            #创建一个名为file_name的excel文件，sheet1为总数据，sheet2为城市数据
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1.append(["name", "total", "return"])
            ws2 = wb.create_sheet("Sheet2")
            ws2.append(["name", "city",  "count"])
            wb.save(file_name)
            print(f"Excel文件{file_name}创建成功")
        return data_list


    def write_to_excel(self, total_data, city_data_list):
        # 打开名为aaa的excel文件，在sheet1中写入total_data，在sheet2中写入city_data
        file_name = self.file_name
        wb = openpyxl.load_workbook(file_name)
        ws1 = wb["Sheet1"]
        ws2 = wb["Sheet2"]
        # 写入数据
        ws1.append(total_data)
        for d in city_data_list:
            ws2.append(d)
        # 保存 Excel 文件
        wb.save(file_name)
        print("数据成功写入 Excel 文件！")



    def remove_old_excel(self):
        for file in os.listdir(self.download_path):
            if file.startswith("账户-投放数据"):
                os.remove(os.path.join(self.download_path, file))

    def is_downloaded(self):
        for file in os.listdir(self.download_path):
            if file.startswith("账户-投放数据"):
                return True
        return False

    # 检查元素是否存在
    def check_element_exists(self, method, condition):
        try:
            el = self.driver.find_element(method,condition)
            return el
        except Exception as e:
            return False

    def download_city_consume_info(self):
        #失败重试10次
        count = 0
        while True:
            try:
                # 删除C:\Users\Administrator\Downloads下以名为“账户-投放数据”开头的文件
                self.remove_old_excel()
                # 点击报表
                report_el = self.check_element_exists(By.XPATH,'//span[text()="报表"]')
                if not report_el:
                    print("没有报表按钮")
                    return 1001,"el_error"
                report_el.click()
                sleep(2)
                #输入开始时间和结束时间
                start_date_input = self.wait.until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="开始时间"]')))
                start_date_input.send_keys(Keys.CONTROL, 'a')
                start_date_input.send_keys(self.start_date)
                end_date_input = self.wait.until(EC.presence_of_element_located((By.XPATH, '//input[@placeholder="截止时间"]')))
                end_date_input.send_keys(Keys.CONTROL, 'a')
                end_date_input.send_keys(self.end_date)
                sleep(1)
                #选取下拉框，下拉框文字为汇总
                self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//div[@class="d-grid d-select-main d-select-main-indicator --color-text-title"]')))[1].click()
                self.wait.until(EC.visibility_of_element_located((By.XPATH, '//span[text()="汇总"]'))).click()
                sleep(1)
                #将滚轮滑动到跳转至
                el_scroll = self.wait.until(EC.presence_of_element_located((By.XPATH, '//div[text()=" 跳转至 "]')))
                self.driver.execute_script("arguments[0].scrollIntoView();", el_scroll)
                self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//div[@class="d-grid d-select-main d-select-main-indicator --color-text-title"]')))[-1].click()
                self.wait.until(EC.visibility_of_element_located((By.XPATH, '//span[text()="城市"]'))).click()
                sleep(1)
                # 点击下载表格
                self.wait.until(EC.presence_of_element_located((By.XPATH, '//span[text()="下载表格"]'))).click()
                sleep(10)
                if self.is_downloaded():
                    print("下载成功")
                    return "success"
            except Exception as e:
                count = count + 1
                print(f"download_city_consume_info - {count}: {e}")
                if count == 5:
                    return 1000,"download_error"
                self.driver.refresh()
                sleep(3)
                continue

    def read_city_consume_info(self,account_name):
        # 读取下载的excel文件
        for file in os.listdir(self.download_path):
            if file.startswith("账户-投放数据"):
                file_name = os.path.join(self.download_path, file)
                # file为csv文件，读取scv数据
                with open(file_name, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    data_list = []
                    for row in reader:
                        # 跳过第一行
                        if row[1] == "城市" or row[1] == "国外城市" or row[1] == "未知" or row[1] =='-':
                            continue
                        data_list.append([account_name,row[1], row[2]])
                    print(f'csv已读取city数据:',data_list)
                    return data_list
        print("未找到文件,需要重新下载")

    def read_csv_row(self, file_name):
        # 读取下载的csv文件
        with open(file_name, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            data_list = []
            for row in reader:
                # 跳过第一行
                if row[1] == "城市":
                    continue
                data_list.append(row)
            print(f'csv已读取city数据')
            return data_list


def main():
    brower = Brower("../excel/xhs.xlsx")
    try:
        brower.handle_alert()
        brower.start()
        print("所有账户为：",brower.items)
    except Exception as e:
        print(e)
    finally:
        brower.quit()


if __name__ == '__main__':
    main()


